"""Excel 导出服务"""

from datetime import datetime
from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlmodel import Session, col, select

from app.models import (
    Project,
    SettlementClaim,
    SettlementPeriod,
    UserDeptLink,
    WorkRecord,
)


def create_export_workbook(
    session: Session,
    dept_id: UUID,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    period_id: Optional[UUID] = None,
    user_id: Optional[UUID] = None,
    project_id: Optional[UUID] = None,
) -> BytesIO:
    """
    创建导出 Excel 工作簿

    Args:
        session: 数据库会话
        dept_id: 部门ID
        start_date: 开始日期 (可选)
        end_date: 结束日期 (可选)
        period_id: 结算周期ID (可选，如果提供则从结算周期获取日期范围)
        user_id: 成员筛选 (可选)
        project_id: 项目筛选 (可选)

    Returns:
        BytesIO: Excel 文件流
    """
    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 获取部门信息
    # dept = session.get(Department, dept_id)

    # 如果提供了结算周期ID，获取日期范围
    if period_id:
        period = session.get(SettlementPeriod, period_id)
        if period:
            start_date = period.start_date
            end_date = period.end_date

    # ==================== Sheet 1: 个人 ====================
    ws1 = wb.active
    assert ws1
    ws1.title = "个人"

    # 表头
    headers1 = [
        "姓名",
        "学号",
        "项目",
        "工作时间",
        "工作内容",
        "详细信息",
        "时长(分钟)",
        "时长(h)",
        "总时长(分钟)",
        "总时长(h)",
        "工资时长(h)",
        "志愿时长(h)",
    ]
    for c, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 获取部门成员
    member_links = session.exec(
        select(UserDeptLink).where(UserDeptLink.dept_id == dept_id)
    ).all()
    if user_id:
        member_links = [link for link in member_links if link.user_id == user_id]

    row = 2
    for link in member_links:
        user = link.user

        # 构建工作记录查询
        record_query = (
            select(WorkRecord)
            .where(WorkRecord.user_id == user.id)
            .where(WorkRecord.dept_id == dept_id)
        )

        if start_date:
            record_query = record_query.where(WorkRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.where(WorkRecord.created_at <= end_date)
        if project_id:
            record_query = record_query.where(WorkRecord.project_id == project_id)

        records = session.exec(
            record_query.order_by(col(WorkRecord.created_at).asc())
        ).all()

        # 计算系统工时
        system_minutes = sum(r.duration_minutes for r in records)
        system_hours = system_minutes / 60

        # 获取申报信息（仅按结算周期导出时有值）
        paid_hours = 0.0
        volunteer_hours = 0.0

        if period_id:
            claim = session.exec(
                select(SettlementClaim)
                .where(SettlementClaim.period_id == period_id)
                .where(SettlementClaim.user_id == user.id)
            ).first()
            if claim:
                paid_hours = claim.paid_hours
                volunteer_hours = claim.volunteer_hours

        # 写入明细（每条记录一行）
        if records:
            for r in records:
                ws1.cell(row=row, column=1, value=user.name).border = thin_border
                ws1.cell(row=row, column=2, value=user.sduid).border = thin_border
                ws1.cell(row=row, column=3, value=r.project.name).border = thin_border
                ws1.cell(
                    row=row, column=4, value=r.created_at.strftime("%Y-%m-%d %H:%M")
                ).border = thin_border
                ws1.cell(row=row, column=5, value=r.description).border = thin_border
                ws1.cell(
                    row=row, column=6, value=r.related_content or "-"
                ).border = thin_border
                ws1.cell(
                    row=row, column=7, value=r.duration_minutes
                ).border = thin_border
                ws1.cell(
                    row=row, column=8, value=round(r.duration_minutes / 60, 2)
                ).border = thin_border
                ws1.cell(row=row, column=9, value=system_minutes).border = thin_border
                ws1.cell(
                    row=row, column=10, value=round(system_hours, 2)
                ).border = thin_border
                ws1.cell(
                    row=row, column=11, value=round(paid_hours, 2)
                ).border = thin_border
                ws1.cell(
                    row=row, column=12, value=round(volunteer_hours, 2)
                ).border = thin_border
                row += 1
        else:
            # 没有明细时保留人员汇总行
            ws1.cell(row=row, column=1, value=user.name).border = thin_border
            ws1.cell(row=row, column=2, value=user.sduid).border = thin_border
            ws1.cell(row=row, column=3, value="-").border = thin_border
            ws1.cell(row=row, column=4, value="-").border = thin_border
            ws1.cell(row=row, column=5, value="-").border = thin_border
            ws1.cell(row=row, column=6, value="-").border = thin_border
            ws1.cell(row=row, column=7, value=0).border = thin_border
            ws1.cell(row=row, column=8, value=0).border = thin_border
            ws1.cell(row=row, column=9, value=system_minutes).border = thin_border
            ws1.cell(
                row=row, column=10, value=round(system_hours, 2)
            ).border = thin_border
            ws1.cell(
                row=row, column=11, value=round(paid_hours, 2)
            ).border = thin_border
            ws1.cell(
                row=row, column=12, value=round(volunteer_hours, 2)
            ).border = thin_border
            row += 1

    # 调整列宽
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 36
    ws1.column_dimensions["F"].width = 28
    ws1.column_dimensions["G"].width = 12
    ws1.column_dimensions["H"].width = 10
    ws1.column_dimensions["I"].width = 14
    ws1.column_dimensions["J"].width = 12
    ws1.column_dimensions["K"].width = 12
    ws1.column_dimensions["L"].width = 12

    # ==================== Sheet 2: 项目 ====================
    ws2 = wb.create_sheet(title="项目")

    # 表头
    headers2 = [
        "项目名",
        "姓名",
        "工作时间",
        "工作内容",
        "详细信息",
        "时长(分钟)",
        "时长(h)",
        "总时长(分钟)",
        "总时长(h)",
    ]
    for c, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 获取项目
    projects = session.exec(select(Project).where(Project.dept_id == dept_id)).all()
    if project_id:
        projects = [p for p in projects if p.id == project_id]

    row = 2
    for project in projects:
        # 构建查询
        record_query = select(WorkRecord).where(WorkRecord.project_id == project.id)

        if start_date:
            record_query = record_query.where(WorkRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.where(WorkRecord.created_at <= end_date)
        if user_id:
            record_query = record_query.where(WorkRecord.user_id == user_id)

        records = session.exec(record_query).all()

        if not records:
            continue

        total_minutes = sum(r.duration_minutes for r in records)
        project_total_hours = round(total_minutes / 60, 2)

        # 每条记录一行，并带项目总时长
        for r in records:
            ws2.cell(row=row, column=1, value=project.name).border = thin_border
            ws2.cell(row=row, column=2, value=r.user.name).border = thin_border
            ws2.cell(
                row=row, column=3, value=r.created_at.strftime("%Y-%m-%d %H:%M")
            ).border = thin_border
            ws2.cell(row=row, column=4, value=r.description).border = thin_border
            ws2.cell(
                row=row, column=5, value=r.related_content or "-"
            ).border = thin_border
            ws2.cell(row=row, column=6, value=r.duration_minutes).border = thin_border
            ws2.cell(
                row=row, column=7, value=round(r.duration_minutes / 60, 2)
            ).border = thin_border
            ws2.cell(row=row, column=8, value=total_minutes).border = thin_border
            ws2.cell(row=row, column=9, value=project_total_hours).border = thin_border
            row += 1

    # 调整列宽
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 36
    ws2.column_dimensions["E"].width = 28
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 14
    ws2.column_dimensions["I"].width = 12

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
