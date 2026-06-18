"""Excel 导出服务"""

from datetime import datetime
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, col, select

from app.core import settings
from app.models import (
    Project,
    SettlementClaim,
    SettlementPeriod,
    SettlementProjectSummary,
    UserDeptLink,
    WorkRecord,
)


def create_export_workbook(
    session: Session,
    dept_id: UUID,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    period_id: UUID | None = None,
    user_id: UUID | None = None,
    project_id: UUID | None = None,
) -> BytesIO:
    """
    创建导出 Excel 工作簿

    Args:
        session: 数据库会话
        dept_id: 部门ID
        start_date: 开始日期 (可选)
        end_date: 结束日期 (可选)
        period_id: 结算周期ID (可选，如果提供则从结算周期获取日期范围)
        user_id: 成员筛选 (可选)
        project_id: 项目筛选 (可选)

    Returns:
        BytesIO: Excel 文件流
    """
    wb = Workbook()
    wb.properties.creator = f"{wb.properties.creator}; {settings.app_name}"
    wb.properties.lastModifiedBy = settings.app_name

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 如果提供了结算周期ID，获取日期范围
    if period_id:
        period = session.get(SettlementPeriod, period_id)
        if period:
            start_date = period.start_date
            end_date = period.end_date

    # ==================== Sheet 1: 个人 ====================
    ws1 = wb.active
    assert ws1
    ws1.title = "个人"

    # 表头
    headers1 = [
        "姓名",
        "项目",
        "工作时间",
        "工作内容",
        "详细信息",
        "时长(h)",
    ]
    for c, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 获取部门成员
    member_links = session.exec(
        select(UserDeptLink).where(UserDeptLink.dept_id == dept_id)
    ).all()
    if user_id:
        member_links = [link for link in member_links if link.user_id == user_id]

    row = 2
    for link in member_links:
        user = link.user

        # 构建工作记录查询
        record_query = (
            select(WorkRecord)
            .where(WorkRecord.user_id == user.id)
            .where(WorkRecord.dept_id == dept_id)
        )

        if start_date:
            record_query = record_query.where(WorkRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.where(WorkRecord.created_at <= end_date)
        if project_id:
            record_query = record_query.where(WorkRecord.project_id == project_id)

        records = session.exec(
            record_query.order_by(col(WorkRecord.created_at).asc())
        ).all()

        # 过滤时长为零的记录
        records = [r for r in records if r.duration_minutes > 0]

        if not records:
            continue

        # 写入明细（每条记录一行）
        for r in records:
            ws1.cell(row=row, column=1, value=user.name).border = thin_border
            ws1.cell(row=row, column=2, value=r.project.name).border = thin_border
            ws1.cell(
                row=row, column=3, value=r.created_at.strftime("%Y-%m-%d %H:%M")
            ).border = thin_border
            ws1.cell(row=row, column=4, value=r.description).border = thin_border
            ws1.cell(
                row=row, column=5, value=r.related_content or "-"
            ).border = thin_border
            ws1.cell(
                row=row, column=6, value=round(r.duration_minutes / 60, 2)
            ).border = thin_border
            row += 1

    sheet1_last_row = row - 1

    # 小时列统一数值格式
    for r in range(2, sheet1_last_row + 1):
        ws1.cell(row=r, column=6).number_format = "0.00"

    # 调整列宽
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 36
    ws1.column_dimensions["E"].width = 28
    ws1.column_dimensions["F"].width = 10

    # ==================== Sheet 2: 项目 ====================
    ws2 = wb.create_sheet(title="项目")

    # 表头
    headers2 = [
        "项目名",
        "姓名",
        "工作时间",
        "工作内容",
        "详细信息",
        "时长(h)",
    ]
    for c, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 获取项目
    projects = session.exec(select(Project).where(Project.dept_id == dept_id)).all()
    if project_id:
        projects = [p for p in projects if p.id == project_id]

    row = 2
    for project in projects:
        # 构建查询
        record_query = select(WorkRecord).where(WorkRecord.project_id == project.id)

        if start_date:
            record_query = record_query.where(WorkRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.where(WorkRecord.created_at <= end_date)
        if user_id:
            record_query = record_query.where(WorkRecord.user_id == user_id)

        records = session.exec(record_query).all()

        # 过滤时长为零的记录
        records = [r for r in records if r.duration_minutes > 0]

        if not records:
            continue

        # 每条记录一行
        for r in records:
            ws2.cell(row=row, column=1, value=project.name).border = thin_border
            ws2.cell(row=row, column=2, value=r.user.name).border = thin_border
            ws2.cell(
                row=row, column=3, value=r.created_at.strftime("%Y-%m-%d %H:%M")
            ).border = thin_border
            ws2.cell(row=row, column=4, value=r.description).border = thin_border
            ws2.cell(
                row=row, column=5, value=r.related_content or "-"
            ).border = thin_border
            ws2.cell(
                row=row, column=6, value=round(r.duration_minutes / 60, 2)
            ).border = thin_border
            row += 1

    sheet2_last_row = row - 1

    for r in range(2, sheet2_last_row + 1):
        ws2.cell(row=r, column=6).number_format = "0.00"

    # 调整列宽
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 36
    ws2.column_dimensions["E"].width = 28
    ws2.column_dimensions["F"].width = 10

    # ==================== Sheet 3: 统计 ====================
    ws3 = wb.create_sheet(title="统计")

    stats_project_query = select(WorkRecord).where(WorkRecord.dept_id == dept_id)
    if start_date:
        stats_project_query = stats_project_query.where(
            WorkRecord.created_at >= start_date
        )
    if end_date:
        stats_project_query = stats_project_query.where(
            WorkRecord.created_at <= end_date
        )
    if user_id:
        stats_project_query = stats_project_query.where(WorkRecord.user_id == user_id)
    if project_id:
        stats_project_query = stats_project_query.where(
            WorkRecord.project_id == project_id
        )

    stats_records = session.exec(stats_project_query).all()

    project_hours: dict[UUID, float] = {}
    for record in stats_records:
        project_hours[record.project_id] = project_hours.get(record.project_id, 0.0) + (
            record.duration_minutes / 60
        )

    stats_projects = []
    if project_hours:
        stats_projects = session.exec(
            select(Project)
            .where(col(Project.id).in_(list(project_hours.keys())))
            .order_by(col(Project.name).asc())
        ).all()

    summary_by_project_id: dict[UUID, SettlementProjectSummary] = {}
    if period_id:
        period_summaries = session.exec(
            select(SettlementProjectSummary).where(
                SettlementProjectSummary.period_id == period_id
            )
        ).all()
        summary_by_project_id = {item.project_id: item for item in period_summaries}

    member_project_hours: dict[tuple[UUID, UUID], float] = {}
    for record in stats_records:
        key = (record.user_id, record.project_id)
        member_project_hours[key] = member_project_hours.get(key, 0.0) + (
            record.duration_minutes / 60
        )

    claim_by_user: dict[UUID, tuple[float, float]] = {}
    if period_id:
        claims = session.exec(
            select(SettlementClaim).where(SettlementClaim.period_id == period_id)
        ).all()
        for claim in claims:
            claim_by_user[claim.user_id] = (claim.volunteer_hours, claim.paid_hours)

    num_projects = len(stats_projects)
    col_name = 1
    col_sduid = 2
    col_first_project = 3
    col_volunteer = col_first_project + num_projects
    col_paid = col_volunteer + 1
    col_total = col_paid + 1

    # 表头行
    header_row = 1
    headers3 = ["姓名", "学号"]
    for project in stats_projects:
        headers3.append(project.name)
    headers3.extend(["志愿时长", "工资时长", "总时长(h)"])

    for c, header in enumerate(headers3, 1):
        cell = ws3.cell(row=header_row, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row = 2
    for link in member_links:
        user = link.user

        row_total = 0.0
        for pi, project in enumerate(stats_projects):
            hours = member_project_hours.get((user.id, project.id), 0.0)
            row_total += hours

        volunteer_hours, paid_hours = claim_by_user.get(user.id, (0.0, 0.0))

        # 跳过时长为零的人
        if row_total == 0 and volunteer_hours == 0 and paid_hours == 0:
            continue

        ws3.cell(row=row, column=col_name, value=user.name).border = thin_border
        ws3.cell(row=row, column=col_sduid, value=user.sduid).border = thin_border

        for pi, project in enumerate(stats_projects):
            hours = member_project_hours.get((user.id, project.id), 0.0)
            ws3.cell(
                row=row, column=col_first_project + pi, value=round(hours, 2)
            ).border = thin_border
            ws3.cell(row=row, column=col_first_project + pi).number_format = "0.00"

        ws3.cell(
            row=row, column=col_volunteer, value=round(volunteer_hours, 2)
        ).border = thin_border
        ws3.cell(row=row, column=col_volunteer).number_format = "0.00"
        ws3.cell(
            row=row, column=col_paid, value=round(paid_hours, 2)
        ).border = thin_border
        ws3.cell(row=row, column=col_paid).number_format = "0.00"
        ws3.cell(
            row=row, column=col_total, value=round(row_total, 2)
        ).border = thin_border
        ws3.cell(row=row, column=col_total).number_format = "0.00"
        row += 1

    total_row = row
    ws3.cell(row=total_row, column=col_sduid, value="合计").border = thin_border
    ws3.cell(row=total_row, column=col_sduid).font = Font(bold=True)

    for pi in range(num_projects):
        cl = get_column_letter(col_first_project + pi)
        ws3.cell(
            row=total_row,
            column=col_first_project + pi,
            value=f"=SUM({cl}{header_row + 1}:{cl}{total_row - 1})",
        ).border = thin_border
        ws3.cell(row=total_row, column=col_first_project + pi).number_format = "0.00"
        ws3.cell(row=total_row, column=col_first_project + pi).font = Font(bold=True)

    for col_idx in (col_volunteer, col_paid, col_total):
        cl = get_column_letter(col_idx)
        ws3.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({cl}{header_row + 1}:{cl}{total_row - 1})",
        ).border = thin_border
        ws3.cell(row=total_row, column=col_idx).number_format = "0.00"
        ws3.cell(row=total_row, column=col_idx).font = Font(bold=True)

    # 项目属性行（表格下方）
    for label, getter in [
        ("完工状态", lambda p: summary_by_project_id.get(p.id)),
        ("项目总结", lambda p: summary_by_project_id.get(p.id)),
    ]:
        attr_row = total_row + 1 if label == "完工状态" else total_row + 2
        ws3.cell(row=attr_row, column=col_sduid, value=label).border = thin_border
        ws3.cell(row=attr_row, column=col_sduid).font = Font(bold=True)
        for pi, project in enumerate(stats_projects):
            summary_row = getter(project)
            if label == "完工状态":
                val = summary_row.status if summary_row else "-"
            else:
                val = summary_row.summary if summary_row else "-"
            ws3.cell(
                row=attr_row, column=col_first_project + pi, value=val
            ).border = thin_border

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 16
    for pi in range(num_projects):
        ws3.column_dimensions[get_column_letter(col_first_project + pi)].width = 16
    ws3.column_dimensions[get_column_letter(col_volunteer)].width = 12
    ws3.column_dimensions[get_column_letter(col_paid)].width = 12
    ws3.column_dimensions[get_column_letter(col_total)].width = 12

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
